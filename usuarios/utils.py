from openpyxl import load_workbook
from .models import Persona, Rol, Usuario
import requests

def process_excel_file(uploaded_file):
    users_created = 0
    wb = load_workbook(uploaded_file)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Verificar si la fila contiene datos válidos
        if not all(row):
            break  # Salir del bucle si la fila está vacía o no contiene datos válidos

        nombre, apellido, identificacion, fecha_nacimiento, correo, username, password, rol_nombre = row

        # Crear la persona
        persona = Persona.objects.create(
            nombre=nombre,
            apellido=apellido,
            identificacion=identificacion,
            fecha_nacimiento=fecha_nacimiento
        )
        
        # Crear o obtener el rol
        rol, _ = Rol.objects.get_or_create(nombre=rol_nombre)
        
        # Crear el usuario
        user = Usuario.objects.create_user(
            username=username,
            email=correo,
            password=password,
            persona=persona,
        )
        
        # Asignar el rol al usuario a través de la tabla intermedia
        rol.usuarios.add(user)
        
        users_created += 1
    
    return users_created


def obtener_imagen_pokemon(numero_pokemon):
    base_url = 'https://pokeapi.co/api/v2/pokemon/{}/'.format(numero_pokemon)
    response = requests.get(base_url)
    
    if response.status_code == 200:
        pokemon_data = response.json()
        sprite_url = pokemon_data['sprites']['front_default']
        return sprite_url
    else:
        return None